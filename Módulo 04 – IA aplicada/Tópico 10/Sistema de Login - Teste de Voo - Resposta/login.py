"""Sistema de login didático com Tkinter e armazenamento em Excel."""

from hashlib import pbkdf2_hmac
from hmac import compare_digest
from pathlib import Path
import secrets
import time
import tkinter as tk
from tkinter import font, messagebox, ttk

from openpyxl import Workbook, load_workbook

ARQUIVO_USUARIOS = Path(__file__).with_name("usuarios.xlsx")
CABECALHOS = ("Nome", "Login", "Senha", "Status", "Nível")
ITERACOES_HASH = 600_000
TENTATIVAS_MAXIMAS = 3
TEMPO_BLOQUEIO = 60
COR_FUNDO, COR_CARD, COR_PRIMARIA = "#EEF1F8", "#FFFFFF", "#4F46E5"
COR_PRIMARIA_HOVER, COR_TEXTO, COR_TEXTO_SECUNDARIO = "#4338CA", "#1F2937", "#6B7280"
COR_ERRO = "#DC2626"


def gerar_hash(senha):
    salt = secrets.token_bytes(16)
    derivado = pbkdf2_hmac("sha256", senha.encode("utf-8"), salt, ITERACOES_HASH)
    return f"pbkdf2_sha256${ITERACOES_HASH}${salt.hex()}${derivado.hex()}"


def verificar_senha(senha, armazenada):
    """Valida hashes novos; texto puro só é aceito durante a migração."""
    if not isinstance(armazenada, str) or not armazenada.startswith("pbkdf2_sha256$"):
        return compare_digest(str(armazenada or ""), senha), False
    try:
        algoritmo, iteracoes, salt_hex, hash_hex = armazenada.split("$", 3)
        if algoritmo != "pbkdf2_sha256":
            return False, True
        derivado = pbkdf2_hmac(
            "sha256", senha.encode("utf-8"), bytes.fromhex(salt_hex), int(iteracoes)
        ).hex()
        return compare_digest(derivado, hash_hex), True
    except (ValueError, TypeError):
        return False, True


def carregar_usuarios():
    if not ARQUIVO_USUARIOS.exists():
        return []
    workbook = load_workbook(ARQUIVO_USUARIOS, read_only=True, data_only=True)
    planilha = workbook.active
    linhas = planilha.iter_rows(values_only=True)
    cabecalhos = [str(valor).strip() if valor is not None else "" for valor in next(linhas)]
    usuarios = []
    for valores in linhas:
        if any(valor is not None and str(valor).strip() for valor in valores):
            usuarios.append(dict(zip(cabecalhos, valores)))
    workbook.close()
    return usuarios


def salvar_usuarios():
    workbook = load_workbook(ARQUIVO_USUARIOS) if ARQUIVO_USUARIOS.exists() else Workbook()
    planilha = workbook.active
    planilha.delete_rows(1, planilha.max_row)
    planilha.append(CABECALHOS)
    for usuario in usuarios:
        planilha.append([usuario.get(cabecalho, "") for cabecalho in CABECALHOS])
    workbook.save(ARQUIVO_USUARIOS)
    workbook.close()


def migrar_senhas():
    alterou = False
    for usuario in usuarios:
        senha = str(usuario.get("Senha", "") or "")
        if senha and not senha.startswith("pbkdf2_sha256$"):
            usuario["Senha"] = gerar_hash(senha)
            alterou = True
    if alterou:
        salvar_usuarios()


def usuario_por_login(login):
    return next((u for u in usuarios if str(u.get("Login", "")).strip() == login), None)


def fechar_dialogo(dialogo):
    dialogo.grab_release()
    dialogo.destroy()


def atualizar_tabela(tabela, contador):
    for item in tabela.get_children():
        tabela.delete(item)
    agora = time.time()
    for indice, usuario in enumerate(usuarios):
        login = str(usuario.get("Login", "")).strip()
        bloqueado = login in bloqueios and bloqueios[login] > agora
        status = "Bloqueado" if bloqueado else str(usuario.get("Status", "")).strip()
        tabela.insert("", "end", iid=str(indice), values=(
            usuario.get("Nome", ""), login, status, usuario.get("Nível", "")
        ), tags=("par" if indice % 2 == 0 else "impar",))
    contador.config(text=f"Total de usuários: {len(usuarios)}")


def dialogo_usuario(indice, tabela, contador):
    existente = usuarios[indice] if indice is not None else None
    dialogo = tk.Toplevel(janela)
    dialogo.title("Editar usuário" if existente else "Cadastrar usuário")
    dialogo.configure(bg=COR_FUNDO)
    dialogo.resizable(False, False)
    dialogo.transient(janela)
    dialogo.grab_set()
    corpo = tk.Frame(dialogo, bg=COR_CARD, padx=24, pady=20)
    corpo.pack(padx=16, pady=16)
    campos = {}
    for chave, rotulo in (("Nome", "Nome"), ("Login", "Login"), ("Senha", "Nova senha")):
        tk.Label(corpo, text=rotulo, bg=COR_CARD, fg=COR_TEXTO, anchor="w").pack(fill="x")
        campo = tk.Entry(corpo, width=34, show="*" if chave == "Senha" else "")
        if existente and chave != "Senha":
            campo.insert(0, str(existente.get(chave, "")))
        campo.pack(pady=(2, 10))
        campos[chave] = campo
    if existente:
        tk.Label(corpo, text="Deixe a senha em branco para mantê-la.", bg=COR_CARD,
                 fg=COR_TEXTO_SECUNDARIO).pack(anchor="w", pady=(0, 8))
    tk.Label(corpo, text="Status", bg=COR_CARD, fg=COR_TEXTO, anchor="w").pack(fill="x")
    status = ttk.Combobox(corpo, values=("Ativo", "Desativado"), state="readonly", width=31)
    status.set(str(existente.get("Status", "Ativo")) if existente else "Ativo")
    status.pack(pady=(2, 10))
    tk.Label(corpo, text="Nível de acesso", bg=COR_CARD, fg=COR_TEXTO, anchor="w").pack(fill="x")
    nivel = ttk.Combobox(corpo, values=("Usuário", "Administrador"), state="readonly", width=31)
    nivel.set(str(existente.get("Nível", "Usuário")) if existente else "Usuário")
    nivel.pack(pady=(2, 14))

    def confirmar():
        nome, login, senha = campos["Nome"].get().strip(), campos["Login"].get().strip(), campos["Senha"].get()
        if not nome or not login or (existente is None and not senha):
            messagebox.showwarning("Dados incompletos", "Preencha nome, login e senha.", parent=dialogo)
            return
        outro = usuario_por_login(login)
        if outro is not None and outro is not existente:
            messagebox.showerror("Login duplicado", "Já existe um usuário com esse login.", parent=dialogo)
            return
        login_anterior = str(existente.get("Login", "")).strip() if existente else ""
        dados = existente if existente else {cabecalho: "" for cabecalho in CABECALHOS}
        dados.update({"Nome": nome, "Login": login, "Status": status.get(), "Nível": nivel.get()})
        if senha:
            dados["Senha"] = gerar_hash(senha)
        if existente and login_anterior != login:
            bloqueios.pop(login_anterior, None)
            tentativas.pop(login_anterior, None)
        if existente is None:
            usuarios.append(dados)
        salvar_usuarios()
        atualizar_tabela(tabela, contador)
        fechar_dialogo(dialogo)

    botoes = tk.Frame(corpo, bg=COR_CARD)
    botoes.pack(fill="x")
    tk.Button(botoes, text="Cancelar", command=lambda: fechar_dialogo(dialogo)).pack(side="right", padx=(8, 0))
    tk.Button(botoes, text="Salvar", bg=COR_PRIMARIA, fg="white", command=confirmar).pack(side="right")


def abrir_painel_administrador():
    painel = tk.Toplevel(janela)
    painel.title("Painel do Administrador")
    painel.configure(bg=COR_FUNDO)
    painel.geometry("760x480")
    painel.transient(janela)
    painel.grab_set()
    tk.Label(painel, text="Usuários cadastrados", font=fonte_titulo_painel,
             bg=COR_FUNDO, fg=COR_TEXTO).pack(anchor="w", padx=24, pady=(20, 10))
    corpo = tk.Frame(painel, bg=COR_FUNDO)
    corpo.pack(fill="both", expand=True, padx=24)
    tabela = ttk.Treeview(corpo, columns=("nome", "login", "status", "nivel"), show="headings")
    for coluna, titulo, largura in (("nome", "Nome", 220), ("login", "Login", 180),
                                    ("status", "Status", 120), ("nivel", "Nível", 160)):
        tabela.heading(coluna, text=titulo)
        tabela.column(coluna, width=largura)
    tabela.pack(side="left", fill="both", expand=True)
    ttk.Scrollbar(corpo, orient="vertical", command=tabela.yview).pack(side="right", fill="y")
    rodape = tk.Frame(painel, bg=COR_FUNDO)
    rodape.pack(fill="x", padx=24, pady=16)
    contador = tk.Label(rodape, bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO)
    contador.pack(side="left")

    def selecionado():
        selecao = tabela.selection()
        return int(selecao[0]) if selecao else None

    def editar():
        indice = selecionado()
        if indice is None:
            messagebox.showwarning("Selecione um usuário", "Selecione um usuário na tabela.", parent=painel)
        else:
            dialogo_usuario(indice, tabela, contador)

    def excluir():
        indice = selecionado()
        if indice is None:
            messagebox.showwarning("Selecione um usuário", "Selecione um usuário na tabela.", parent=painel)
            return
        if usuarios[indice] is usuario_atual:
            messagebox.showwarning("Operação não permitida", "O administrador conectado não pode ser excluído.", parent=painel)
            return
        if messagebox.askyesno("Confirmar exclusão", "Deseja realmente excluir este usuário?", parent=painel):
            login = str(usuarios[indice].get("Login", "")).strip()
            usuarios.pop(indice)
            bloqueios.pop(login, None)
            tentativas.pop(login, None)
            salvar_usuarios()
            atualizar_tabela(tabela, contador)

    def desbloquear():
        indice = selecionado()
        if indice is None:
            messagebox.showwarning("Selecione um usuário", "Selecione um usuário bloqueado.", parent=painel)
            return
        login = str(usuarios[indice].get("Login", "")).strip()
        bloqueios.pop(login, None)
        tentativas.pop(login, None)
        atualizar_tabela(tabela, contador)

    botoes = tk.Frame(painel, bg=COR_FUNDO)
    botoes.pack(fill="x", padx=24, pady=(0, 20))
    for texto, comando in (("Cadastrar", lambda: dialogo_usuario(None, tabela, contador)),
                           ("Editar", editar), ("Excluir", excluir), ("Desbloquear", desbloquear)):
        tk.Button(botoes, text=texto, command=comando).pack(side="left", padx=(0, 8))
    tk.Button(botoes, text="Fechar", command=lambda: fechar_dialogo(painel)).pack(side="right")
    atualizar_tabela(tabela, contador)


def exibir_erro(mensagem):
    label_erro.config(text=mensagem)


def realizar_login():
    global usuario_atual
    login, senha = campo_login.get().strip(), campo_senha.get()
    exibir_erro("")
    if not login or not senha:
        exibir_erro("Preencha login e senha.")
        return
    agora = time.time()
    if login in bloqueios and bloqueios[login] > agora:
        restante = int(bloqueios[login] - agora) + 1
        exibir_erro(f"Login bloqueado. Tente novamente em {restante}s ou peça o desbloqueio ao administrador.")
        return
    bloqueios.pop(login, None)
    usuario = usuario_por_login(login)
    valido, ja_hashed = verificar_senha(senha, usuario.get("Senha", "")) if usuario else (False, True)
    if not usuario or not valido:
        tentativas[login] = tentativas.get(login, 0) + 1
        if tentativas[login] >= TENTATIVAS_MAXIMAS:
            bloqueios[login] = agora + TEMPO_BLOQUEIO
            exibir_erro("Muitas tentativas incorretas. Login bloqueado por 60 segundos.")
        else:
            exibir_erro("Login ou senha incorretos.")
        return
    tentativas.pop(login, None)
    if not ja_hashed:
        usuario["Senha"] = gerar_hash(senha)
        salvar_usuarios()
    if str(usuario.get("Status", "")).strip().casefold() != "ativo":
        exibir_erro(f"Usuário desativado. Procure seu líder. (Nome: {usuario.get('Nome', '')})")
        return
    usuario_atual = usuario
    messagebox.showinfo("Login bem-sucedido", f"Login bem-sucedido!\nNome: {usuario.get('Nome', 'Usuário')}")
    if str(usuario.get("Nível", "")).strip().casefold() == "administrador":
        abrir_painel_administrador()


def alternar_senha():
    visivel = campo_senha.cget("show") == ""
    campo_senha.config(show="*" if visivel else "")
    botao_olho.config(text="👁" if visivel else "🙈")


def centralizar_janela(root, largura, altura):
    root.update_idletasks()
    root.geometry(f"{largura}x{altura}+{(root.winfo_screenwidth()-largura)//2}+{(root.winfo_screenheight()-altura)//2}")


usuarios = carregar_usuarios()
tentativas, bloqueios, usuario_atual = {}, {}, None
migrar_senhas()

janela = tk.Tk()
janela.title("Sistema de Login")
janela.configure(bg=COR_FUNDO)
janela.resizable(False, False)
centralizar_janela(janela, 420, 430)
fonte_titulo = font.Font(family="Segoe UI", size=19, weight="bold")
fonte_subtitulo = font.Font(family="Segoe UI", size=10)
fonte_label = font.Font(family="Segoe UI", size=10, weight="bold")
fonte_entry = font.Font(family="Segoe UI", size=11)
fonte_botao = font.Font(family="Segoe UI", size=11, weight="bold")
fonte_erro = font.Font(family="Segoe UI", size=9)
fonte_titulo_painel = font.Font(family="Segoe UI", size=15, weight="bold")

tk.Label(janela, text="🔐", font=("Segoe UI Emoji", 28), bg=COR_PRIMARIA, fg="white").pack(fill="x", pady=(14, 8))
card = tk.Frame(janela, bg=COR_CARD, padx=34, pady=24)
card.pack(fill="x", padx=20)
tk.Label(card, text="Acesso ao sistema", font=fonte_titulo, bg=COR_CARD, fg=COR_TEXTO).pack()
tk.Label(card, text="Entre com suas credenciais para continuar", font=fonte_subtitulo,
         bg=COR_CARD, fg=COR_TEXTO_SECUNDARIO).pack(pady=(2, 20))
tk.Label(card, text="LOGIN", font=fonte_label, bg=COR_CARD, fg=COR_TEXTO_SECUNDARIO).pack(anchor="w")
campo_login = tk.Entry(card, font=fonte_entry, relief="solid", bd=1)
campo_login.pack(fill="x", pady=(4, 14), ipady=6)
tk.Label(card, text="SENHA", font=fonte_label, bg=COR_CARD, fg=COR_TEXTO_SECUNDARIO).pack(anchor="w")
linha_senha = tk.Frame(card, bg=COR_CARD)
linha_senha.pack(fill="x", pady=(4, 4))
campo_senha = tk.Entry(linha_senha, font=fonte_entry, show="*", relief="solid", bd=1)
campo_senha.pack(side="left", fill="x", expand=True, ipady=6)
botao_olho = tk.Label(linha_senha, text="👁", font=("Segoe UI Emoji", 10), bg=COR_CARD,
                      fg=COR_TEXTO_SECUNDARIO, cursor="hand2")
botao_olho.pack(side="right", padx=8)
botao_olho.bind("<Button-1>", lambda _evento: alternar_senha())
label_erro = tk.Label(card, text="", font=fonte_erro, bg=COR_CARD, fg=COR_ERRO, anchor="w")
label_erro.pack(fill="x", pady=(2, 4))
botao_entrar = tk.Button(card, text="Entrar", font=fonte_botao, bg=COR_PRIMARIA, fg="white",
                         activebackground=COR_PRIMARIA_HOVER, activeforeground="white", bd=0,
                         command=realizar_login)
botao_entrar.pack(fill="x", ipady=8, pady=(8, 0))
tk.Label(janela, text="© Sistema de Login", font=("Segoe UI", 8), bg=COR_FUNDO,
         fg=COR_TEXTO_SECUNDARIO).pack(pady=10)
campo_login.focus()
janela.bind("<Return>", lambda _evento: realizar_login())
janela.mainloop()
