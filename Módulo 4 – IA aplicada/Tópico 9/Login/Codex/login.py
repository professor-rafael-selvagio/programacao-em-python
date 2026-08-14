"""
Sistema de login com Tkinter
e dados armazenados em uma
planilha Excel.

Versão com melhorias de design (cores, tipografia, layout em "card",
hover no botão, mostrar/ocultar senha, janela centralizada).
"""

from pathlib import Path
import tkinter as tk
from tkinter import messagebox, font, ttk

from openpyxl import load_workbook


ARQUIVO_USUARIOS = Path(__file__).with_name("usuarios.xlsx")

# ---------------------------------------------------------------------------
# Paleta de cores
# ---------------------------------------------------------------------------
COR_FUNDO = "#EEF1F8"          # fundo da janela
COR_CARD = "#FFFFFF"           # cartão central
COR_PRIMARIA = "#4F46E5"       # roxo/azul (botão, título)
COR_PRIMARIA_HOVER = "#4338CA"
COR_TEXTO = "#1F2937"
COR_TEXTO_SECUNDARIO = "#6B7280"
COR_BORDA = "#D1D5DB"
COR_BORDA_FOCO = "#4F46E5"
COR_ERRO = "#DC2626"


def carregar_usuarios():
    """Lê os usuários da planilha e retorna uma lista de dicionários."""
    workbook = load_workbook(ARQUIVO_USUARIOS, read_only=True, data_only=True)
    planilha = workbook.active

    linhas = planilha.iter_rows(values_only=True)
    cabecalhos = [str(valor).strip() if valor is not None else "" for valor in next(linhas)]
    usuarios = []

    for valores in linhas:
        usuario = dict(zip(cabecalhos, valores))
        usuarios.append(usuario)

    workbook.close()
    return usuarios


def realizar_login():
    """Valida o login, a senha e o status do usuário."""
    login_informado = campo_login.get().strip()
    senha_informada = campo_senha.get()

    limpar_erro()

    if not login_informado or not senha_informada:
        mostrar_erro("Preencha login e senha.")
        return

    usuario_encontrado = next(
        (
            usuario
            for usuario in usuarios
            if str(usuario.get("Login", "")).strip() == login_informado
            and str(usuario.get("Senha", "")) == senha_informada
        ),
        None,
    )

    if usuario_encontrado is None:
        mostrar_erro("Login ou senha incorretos.")
        return

    nome = usuario_encontrado.get("Nome", "Usuário")
    status = str(usuario_encontrado.get("Status", "")).strip().casefold()
    nivel = str(usuario_encontrado.get("Nível", "")).strip().casefold()

    if status != "ativo":
        mostrar_erro(f"Usuário desativado. Procure seu líder. (Nome: {nome})")
        return

    messagebox.showinfo("Login bem-sucedido", f"Login bem-sucedido!\nNome: {nome}")

    if nivel == "administrador":
        abrir_painel_administrador()


def abrir_painel_administrador():
    """Abre uma janela listando todos os usuários cadastrados."""
    painel = tk.Toplevel(janela)
    painel.title("Painel do Administrador")
    painel.configure(bg=COR_FUNDO)

    LARGURA_PAINEL, ALTURA_PAINEL = 720, 460
    painel.geometry(f"{LARGURA_PAINEL}x{ALTURA_PAINEL}")
    painel.minsize(560, 340)
    painel.transient(janela)
    painel.grab_set()

    # Centraliza o painel em relação à janela principal
    janela.update_idletasks()
    x = janela.winfo_x() + (janela.winfo_width() // 2) - (LARGURA_PAINEL // 2)
    y = janela.winfo_y() + (janela.winfo_height() // 2) - (ALTURA_PAINEL // 2)
    painel.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    faixa = tk.Frame(painel, bg=COR_PRIMARIA, height=70)
    faixa.pack(fill="x")

    tk.Label(
        faixa, text="👥  Usuários cadastrados", font=fonte_titulo_painel,
        bg=COR_PRIMARIA, fg="white",
    ).pack(side="left", padx=24, pady=16)

    corpo = tk.Frame(painel, bg=COR_FUNDO)
    corpo.pack(fill="both", expand=True, padx=20, pady=16)

    estilo = ttk.Style(painel)
    estilo.theme_use("clam")
    estilo.configure(
        "Admin.Treeview",
        background=COR_CARD,
        fieldbackground=COR_CARD,
        foreground=COR_TEXTO,
        rowheight=28,
        font=fonte_entry,
        borderwidth=0,
    )
    estilo.configure(
        "Admin.Treeview.Heading",
        background=COR_PRIMARIA,
        foreground="white",
        font=fonte_label,
        relief="flat",
    )
    estilo.map("Admin.Treeview.Heading", background=[("active", COR_PRIMARIA_HOVER)])
    estilo.map(
        "Admin.Treeview",
        background=[("selected", COR_PRIMARIA)],
        foreground=[("selected", "white")],
    )

    colunas = ("nome", "login", "senha", "status")
    tabela = ttk.Treeview(
        corpo, columns=colunas, show="headings", style="Admin.Treeview",
    )
    tabela.heading("nome", text="Nome")
    tabela.heading("login", text="Login")
    tabela.heading("senha", text="Senha")
    tabela.heading("status", text="Status")

    tabela.column("nome", width=220, anchor="w")
    tabela.column("login", width=140, anchor="w")
    tabela.column("senha", width=120, anchor="w")
    tabela.column("status", width=100, anchor="center")

    barra_rolagem = ttk.Scrollbar(corpo, orient="vertical", command=tabela.yview)
    tabela.configure(yscrollcommand=barra_rolagem.set)

    tabela.pack(side="left", fill="both", expand=True)
    barra_rolagem.pack(side="right", fill="y")

    for indice, usuario in enumerate(usuarios):
        status_usuario = str(usuario.get("Status", "")).strip()
        tag = "par" if indice % 2 == 0 else "impar"
        tabela.insert(
            "",
            "end",
            values=(
                usuario.get("Nome", ""),
                usuario.get("Login", ""),
                usuario.get("Senha", ""),
                status_usuario,
            ),
            tags=(tag,),
        )

    tabela.tag_configure("par", background=COR_CARD)
    tabela.tag_configure("impar", background="#F4F6FB")

    rodape_painel = tk.Frame(painel, bg=COR_FUNDO)
    rodape_painel.pack(fill="x", padx=20, pady=(0, 16))

    tk.Label(
        rodape_painel, text=f"Total de usuários: {len(usuarios)}",
        font=fonte_erro, bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO,
    ).pack(side="left")

    botao_fechar = tk.Button(
        rodape_painel, text="Fechar", font=fonte_label, bg=COR_PRIMARIA, fg="white",
        activebackground=COR_PRIMARIA_HOVER, activeforeground="white", bd=0,
        relief="flat", cursor="hand2", command=painel.destroy,
    )
    botao_fechar.pack(side="right", ipadx=14, ipady=4)
    botao_fechar.bind("<Enter>", lambda _e: botao_fechar.config(bg=COR_PRIMARIA_HOVER))
    botao_fechar.bind("<Leave>", lambda _e: botao_fechar.config(bg=COR_PRIMARIA))


def mostrar_erro(mensagem):
    label_erro.config(text=mensagem)


def limpar_erro():
    label_erro.config(text="")


def alternar_senha():
    """Mostra/oculta o texto do campo de senha."""
    if campo_senha.cget("show") == "*":
        campo_senha.config(show="")
        botao_olho.config(text="🙈")
    else:
        campo_senha.config(show="*")
        botao_olho.config(text="👁")


def estilizar_entry_foco(widget, evento_borda):
    """Aplica um contorno colorido quando o campo recebe foco."""
    def ao_focar(_evento):
        evento_borda.config(highlightbackground=COR_BORDA_FOCO, highlightcolor=COR_BORDA_FOCO)

    def ao_sair(_evento):
        evento_borda.config(highlightbackground=COR_BORDA, highlightcolor=COR_BORDA)

    widget.bind("<FocusIn>", ao_focar)
    widget.bind("<FocusOut>", ao_sair)


def centralizar_janela(root, largura, altura):
    root.update_idletasks()
    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()
    x = (largura_tela // 2) - (largura // 2)
    y = (altura_tela // 2) - (altura // 2)
    root.geometry(f"{largura}x{altura}+{x}+{y}")


try:
    usuarios = carregar_usuarios()
except (FileNotFoundError, KeyError, StopIteration, OSError) as erro:
    usuarios = []
    erro_planilha = str(erro)
else:
    erro_planilha = ""


# ---------------------------------------------------------------------------
# Janela principal
# ---------------------------------------------------------------------------
janela = tk.Tk()
janela.title("Sistema de Login")
janela.configure(bg=COR_FUNDO)
janela.resizable(False, False)

LARGURA, ALTURA = 420, 480
centralizar_janela(janela, LARGURA, ALTURA)

fonte_titulo = font.Font(family="Segoe UI", size=19, weight="bold")
fonte_subtitulo = font.Font(family="Segoe UI", size=10)
fonte_label = font.Font(family="Segoe UI", size=10, weight="bold")
fonte_entry = font.Font(family="Segoe UI", size=11)
fonte_botao = font.Font(family="Segoe UI", size=11, weight="bold")
fonte_erro = font.Font(family="Segoe UI", size=9)
fonte_titulo_painel = font.Font(family="Segoe UI", size=15, weight="bold")

# Faixa superior colorida (puramente decorativa)
faixa_topo = tk.Frame(janela, bg=COR_PRIMARIA, height=90)
faixa_topo.pack(fill="x")

icone_usuario = tk.Label(faixa_topo, text="🔐", font=("Segoe UI Emoji", 28), bg=COR_PRIMARIA)
icone_usuario.pack(pady=(18, 0))

# Cartão central (com pequena margem para simular sombra/elevação)
sombra = tk.Frame(janela, bg="#C7CCDA")
sombra.place(x=24, y=70, width=LARGURA - 44, height=ALTURA - 100)

card = tk.Frame(janela, bg=COR_CARD)
card.place(x=20, y=64, width=LARGURA - 44, height=ALTURA - 100)

titulo = tk.Label(card, text="Acesso ao sistema", font=fonte_titulo, bg=COR_CARD, fg=COR_TEXTO)
titulo.pack(pady=(26, 2))

subtitulo = tk.Label(
    card,
    text="Entre com suas credenciais para continuar",
    font=fonte_subtitulo,
    bg=COR_CARD,
    fg=COR_TEXTO_SECUNDARIO,
)
subtitulo.pack(pady=(0, 22))

conteudo = tk.Frame(card, bg=COR_CARD)
conteudo.pack(fill="x", padx=34)

# --- Campo Login ---
tk.Label(conteudo, text="LOGIN", font=fonte_label, bg=COR_CARD, fg=COR_TEXTO_SECUNDARIO).pack(
    anchor="w"
)

borda_login = tk.Frame(
    conteudo, bg=COR_CARD, highlightbackground=COR_BORDA, highlightcolor=COR_BORDA,
    highlightthickness=1, bd=0,
)
borda_login.pack(fill="x", pady=(4, 16))

campo_login = tk.Entry(
    borda_login, font=fonte_entry, bd=0, bg=COR_CARD, fg=COR_TEXTO,
    insertbackground=COR_TEXTO, relief="flat",
)
campo_login.pack(fill="x", padx=10, pady=8)
estilizar_entry_foco(campo_login, borda_login)

# --- Campo Senha ---
tk.Label(conteudo, text="SENHA", font=fonte_label, bg=COR_CARD, fg=COR_TEXTO_SECUNDARIO).pack(
    anchor="w"
)

borda_senha = tk.Frame(
    conteudo, bg=COR_CARD, highlightbackground=COR_BORDA, highlightcolor=COR_BORDA,
    highlightthickness=1, bd=0,
)
borda_senha.pack(fill="x", pady=(4, 4))

campo_senha = tk.Entry(
    borda_senha, font=fonte_entry, bd=0, bg=COR_CARD, fg=COR_TEXTO,
    insertbackground=COR_TEXTO, relief="flat", show="*",
)
campo_senha.pack(side="left", fill="x", expand=True, padx=(10, 0), pady=8)
estilizar_entry_foco(campo_senha, borda_senha)

botao_olho = tk.Label(
    borda_senha, text="👁", font=("Segoe UI Emoji", 10), bg=COR_CARD,
    fg=COR_TEXTO_SECUNDARIO, cursor="hand2",
)
botao_olho.pack(side="right", padx=10)
botao_olho.bind("<Button-1>", lambda _evento: alternar_senha())

# --- Mensagem de erro (inline) ---
label_erro = tk.Label(conteudo, text="", font=fonte_erro, bg=COR_CARD, fg=COR_ERRO, anchor="w")
label_erro.pack(fill="x", pady=(2, 4))

# --- Botão Entrar (com efeito hover) ---
botao_entrar = tk.Button(
    conteudo,
    text="Entrar",
    font=fonte_botao,
    bg=COR_PRIMARIA,
    fg="white",
    activebackground=COR_PRIMARIA_HOVER,
    activeforeground="white",
    bd=0,
    relief="flat",
    cursor="hand2",
    command=realizar_login,
)
botao_entrar.pack(fill="x", ipady=9, pady=(10, 0))

botao_entrar.bind("<Enter>", lambda _e: botao_entrar.config(bg=COR_PRIMARIA_HOVER))
botao_entrar.bind("<Leave>", lambda _e: botao_entrar.config(bg=COR_PRIMARIA))

rodape = tk.Label(
    janela, text="© Sistema de Login", font=("Segoe UI", 8), bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO
)
rodape.pack(side="bottom", pady=8)

if erro_planilha:
    messagebox.showerror("Erro", f"Não foi possível carregar a planilha:\n{erro_planilha}")

janela.bind("<Return>", lambda evento: realizar_login())
campo_login.focus()
janela.mainloop()