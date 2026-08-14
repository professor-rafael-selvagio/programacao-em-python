from pathlib import Path
from openpyxl import Workbook, load_workbook


ARQUIVO_USUARIOS = Path(__file__).resolve().parent / "usuarios.xlsx"

COLUNAS_OBRIGATORIAS = [
    "Nome",
    "Login",
    "Senha",
    "Status",
    "Nível"
]


def criar_planilha():
    """
    Cria a planilha inicial caso ela ainda não exista.
    """

    if ARQUIVO_USUARIOS.exists():
        return

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Usuarios"

    worksheet.append(COLUNAS_OBRIGATORIAS)

    usuarios = [
        ["João Silva", "joao", "1234", "Ativo", "Usuário"],
        ["Maria Souza", "maria", "5678", "Ativo", "Administrador"],
        ["Pedro Santos", "pedro", "1111", "Inativo", "Usuário"],
        ["Ana Lima", "ana", "2222", "Ativo", "Usuário"],
        ["Carlos Oliveira", "carlos", "3333", "Inativo", "Administrador"],
        ["Fernanda Costa", "fernanda", "4444", "Ativo", "Administrador"]
    ]

    for usuario in usuarios:
        worksheet.append(usuario)

    # Ajusta a largura das colunas
    larguras = {
        "A": 25,
        "B": 18,
        "C": 18,
        "D": 15,
        "E": 20
    }

    for coluna, largura in larguras.items():
        worksheet.column_dimensions[coluna].width = largura

    # Deixa o cabeçalho em negrito
    for celula in worksheet[1]:
        celula.font = celula.font.copy(bold=True)

    worksheet.freeze_panes = "A2"

    workbook.save(ARQUIVO_USUARIOS)


def carregar_usuarios():
    """
    Carrega os usuários da planilha.

    Retorna:
        list: lista de dicionários contendo os usuários.

    Levanta:
        Exception: quando existe algum problema com a planilha.
    """

    criar_planilha()

    try:
        workbook = load_workbook(
            ARQUIVO_USUARIOS,
            data_only=True
        )

        worksheet = workbook.active

        linhas = list(
            worksheet.iter_rows(values_only=True)
        )

        if not linhas:
            raise Exception(
                "A planilha está vazia."
            )

        cabecalho = [
            str(valor).strip()
            if valor is not None
            else ""
            for valor in linhas[0]
        ]

        # Verifica se todas as colunas existem
        colunas_ausentes = [
            coluna
            for coluna in COLUNAS_OBRIGATORIAS
            if coluna not in cabecalho
        ]

        if colunas_ausentes:
            raise Exception(
                "A planilha não possui as colunas obrigatórias: "
                + ", ".join(colunas_ausentes)
            )

        indices = {
            coluna: cabecalho.index(coluna)
            for coluna in COLUNAS_OBRIGATORIAS
        }

        usuarios = []

        for linha in linhas[1:]:

            if not linha:
                continue

            if all(valor is None for valor in linha):
                continue

            usuario = {}

            for coluna, indice in indices.items():

                if indice < len(linha):
                    valor = linha[indice]
                else:
                    valor = ""

                usuario[coluna] = (
                    str(valor).strip()
                    if valor is not None
                    else ""
                )

            usuarios.append(usuario)

        return usuarios

    except Exception as erro:

        raise Exception(
            f"Não foi possível carregar a planilha.\n\n"
            f"Detalhes: {erro}"
        )


def autenticar(login, senha):
    """
    Realiza a autenticação do usuário.

    A validação acontece nesta ordem:

    1. Login e senha
    2. Status
    3. Acesso autorizado
    """

    usuarios = carregar_usuarios()

    usuario_encontrado = None

    # 1. Verifica login e senha
    for usuario in usuarios:

        if (
            usuario["Login"] == login
            and usuario["Senha"] == senha
        ):
            usuario_encontrado = usuario
            break

    # Login ou senha incorretos
    if usuario_encontrado is None:

        return {
            "sucesso": False,
            "tipo": "credenciais",
            "mensagem": "Login ou senha incorretos."
        }

    # 2. Verifica o status
    if usuario_encontrado["Status"].lower() != "ativo":

        return {
            "sucesso": False,
            "tipo": "inativo",
            "mensagem": (
                "Este usuário está inativo. "
                "Entre em contato com o administrador."
            )
        }

    # 3. Login autorizado
    return {
        "sucesso": True,
        "tipo": "sucesso",
        "mensagem": (
            f"Bem-vindo, {usuario_encontrado['Nome']}!"
        ),
        "usuario": usuario_encontrado
    }


def usuario_e_administrador(usuario):
    """
    Verifica se o usuário possui nível administrativo.
    """

    return (
        usuario is not None
        and usuario["Nível"].lower()
        == "administrador"
    )