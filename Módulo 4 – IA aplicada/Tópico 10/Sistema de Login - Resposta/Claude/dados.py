"""
dados.py
--------
Camada de dados (backend) do sistema de login.

Responsável por:
    - Carregar os usuários a partir da planilha 'usuarios.xlsx';
    - Validar credenciais (login/senha);
    - Verificar status (Ativo/Inativo) e nível (Usuário/Administrador);
    - Tratar erros relacionados ao arquivo (ausente, corrompido, mal formatado).

Nenhuma função aqui depende do Tkinter: essa separação permite testar
a lógica de autenticação de forma isolada da interface gráfica.
"""

import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

PASTA_PROJETO = os.path.dirname(os.path.abspath(__file__))
NOME_ARQUIVO = os.path.join(PASTA_PROJETO, "usuarios.xlsx")

# Colunas esperadas na planilha, na ordem correta.
COLUNAS_ESPERADAS = ["Nome", "Login", "Senha", "Status", "Nível"]


class ErroCarregarPlanilha(Exception):
    """
    Exceção customizada usada para sinalizar problemas ao carregar a
    planilha de usuários (arquivo ausente, corrompido, vazio ou com
    colunas incorretas). A interface gráfica captura essa exceção e
    exibe uma mensagem amigável ao usuário.
    """
    pass


def carregar_usuarios():
    """
    Lê a planilha 'usuarios.xlsx' e retorna uma lista de dicionários,
    um para cada usuário cadastrado, no formato:

        {
            "Nome": str,
            "Login": str,
            "Senha": str,
            "Status": str,   # "Ativo" ou "Inativo"
            "Nível": str,    # "Usuário" ou "Administrador"
        }

    Lança ErroCarregarPlanilha em caso de qualquer problema na leitura.
    """
    # 1) Verifica se o arquivo existe antes de tentar abrir.
    if not os.path.exists(NOME_ARQUIVO):
        raise ErroCarregarPlanilha(
            f"Arquivo '{NOME_ARQUIVO}' não encontrado.\n"
            "Execute 'gerar_planilha.py' para criá-lo ou verifique o local do arquivo."
        )

    # 2) Tenta abrir o arquivo; captura corrupção/formato inválido.
    try:
        workbook = load_workbook(NOME_ARQUIVO, data_only=True)
        planilha = workbook.active
    except InvalidFileException:
        raise ErroCarregarPlanilha(
            f"O arquivo '{NOME_ARQUIVO}' não é um arquivo Excel válido (.xlsx)."
        )
    except Exception as erro:
        # Cobre outros problemas de leitura (arquivo corrompido, permissão, etc.)
        raise ErroCarregarPlanilha(
            f"Não foi possível abrir '{NOME_ARQUIVO}'.\nDetalhes: {erro}"
        )

    linhas = list(planilha.iter_rows(values_only=True))

    # 3) Planilha vazia (nem cabeçalho).
    if not linhas:
        raise ErroCarregarPlanilha(f"A planilha '{NOME_ARQUIVO}' está vazia.")

    cabecalho = list(linhas[0])

    # 4) Verifica se todas as colunas esperadas estão presentes.
    if cabecalho != COLUNAS_ESPERADAS:
        raise ErroCarregarPlanilha(
            "A planilha não possui as colunas esperadas.\n"
            f"Esperado: {COLUNAS_ESPERADAS}\n"
            f"Encontrado: {cabecalho}"
        )

    linhas_dados = linhas[1:]

    # 5) Planilha só com cabeçalho, sem nenhum usuário cadastrado.
    if not linhas_dados:
        raise ErroCarregarPlanilha("A planilha não contém nenhum usuário cadastrado.")

    usuarios = []
    for indice, linha in enumerate(linhas_dados, start=2):
        # Ignora linhas completamente vazias (ex: linhas em branco no final).
        if linha is None or all(valor is None for valor in linha):
            continue

        # Linha incompleta (faltando alguma coluna) é tratada como erro,
        # pois pode indicar planilha corrompida ou editada incorretamente.
        if len(linha) < len(COLUNAS_ESPERADAS) or any(valor is None for valor in linha[:5]):
            raise ErroCarregarPlanilha(
                f"Linha {indice} da planilha está incompleta ou mal formatada."
            )

        nome, login, senha, status, nivel = linha[:5]
        usuarios.append({
            "Nome": str(nome).strip(),
            "Login": str(login).strip(),
            "Senha": str(senha).strip(),
            "Status": str(status).strip(),
            "Nível": str(nivel).strip(),
        })

    return usuarios


def autenticar(login, senha, usuarios):
    """
    Valida as credenciais informadas contra a lista de usuários carregada.

    Retorna uma tupla (resultado, usuario):
        - ("vazio", None)        -> login ou senha não preenchidos
        - ("invalido", None)     -> login não existe OU senha incorreta
        - ("inativo", usuario)   -> credenciais corretas, mas usuário inativo
        - ("sucesso", usuario)   -> autenticação bem-sucedida

    A ordem de checagem segue exatamente o que foi especificado:
        1) campos vazios
        2) login/senha incorretos (mensagem genérica)
        3) usuário inativo (mensagem específica)
    """
    # 1) Campos vazios (após remover espaços em branco nas pontas).
    if not login.strip() or not senha.strip():
        return "vazio", None

    login = login.strip()
    senha = senha.strip()

    # Procura um usuário cujo login e senha batam exatamente.
    usuario_encontrado = None
    for usuario in usuarios:
        if usuario["Login"] == login and usuario["Senha"] == senha:
            usuario_encontrado = usuario
            break

    # 2) Login não encontrado ou senha incorreta -> mensagem genérica,
    #    sem revelar qual dos dois campos está errado (por segurança).
    if usuario_encontrado is None:
        return "invalido", None

    # 3) Usuário e senha corretos, mas conta inativa.
    if usuario_encontrado["Status"].lower() != "ativo":
        return "inativo", usuario_encontrado

    # 4) Tudo certo: autenticação bem-sucedida.
    return "sucesso", usuario_encontrado


def eh_administrador(usuario):
    """Retorna True se o usuário autenticado possui nível de Administrador."""
    return usuario is not None and usuario["Nível"].strip().lower() == "administrador"